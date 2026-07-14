from pathlib import Path

import pytest

from table_extract.profiling import profile_file_context
from table_extract.runtime import FileScanContext, StoredFile
from table_extract.runtime.listener import default_file_scan_callback
from table_extract.sources import (
    DatabaseScanRequest,
    SourceAdapter,
    build_file_source_adapter,
)


def stored_file(path: Path, *, extension: str, mime_type: str) -> StoredFile:
    return StoredFile(
        file_id="file-001",
        run_id="run-001",
        source_type="local",
        source_uri=str(path),
        external_id=None,
        file_name=path.name,
        relative_path=path.name,
        extension=extension,
        mime_type=mime_type,
        size_bytes=path.stat().st_size,
        checksum_sha256=None,
    )


def file_context(path: Path, *, extension: str, mime_type: str) -> FileScanContext:
    file = stored_file(path, extension=extension, mime_type=mime_type)
    return FileScanContext(
        run_id=file.run_id,
        stored_file=file,
        local_path=str(path),
        source_uri=file.source_uri,
        is_temporary=False,
    )


def test_profile_file_context_profiles_csv_structure_and_row_count(tmp_path: Path) -> None:
    csv_path = tmp_path / "clientes.csv"
    csv_path.write_text(
        "email,,rut\n"
        "a@example.com,,12.345.678-5\n"
        "\n"
        "b@example.com,,11.111.111-1\n",
        encoding="utf-8",
    )
    context = file_context(csv_path, extension=".csv", mime_type="text/csv")

    profile = profile_file_context(context)

    assert profile.source_name == "clientes.csv"
    assert profile.source_type == "csv"
    assert profile.source_uri == str(csv_path)
    assert len(profile.tables) == 1
    table = profile.tables[0]
    assert table.table_name == "clientes.csv"
    assert table.table_type == "csv"
    assert table.row_count == 2
    assert [(column.column_name, column.ordinal_position, column.data_type) for column in table.columns] == [
        ("email", 1, "unknown"),
        ("column_2", 2, "unknown"),
        ("rut", 3, "unknown"),
    ]


def test_csv_file_adapter_samples_column_values(tmp_path: Path) -> None:
    csv_path = tmp_path / "clientes.csv"
    csv_path.write_text(
        "email,rut\n"
        "  a@example.com  ,12.345.678-5\n"
        ",\n"
        "   ,11.111.111-1\n"
        "b@example.com,22.222.222-2\n"
        "largo@example.com,33.333.333-3\n",
        encoding="utf-8",
    )
    context = file_context(csv_path, extension=".csv", mime_type="text/csv")
    adapter = build_file_source_adapter(context)

    assert isinstance(adapter, SourceAdapter)

    table = next(iter(adapter.iter_tables()))
    column = next(iter(adapter.iter_columns(table)))

    sample = adapter.get_column_sample(table, column, limit=2, max_value_length=10)

    assert sample.table_name == "clientes.csv"
    assert sample.column_name == "email"
    assert sample.values == ("a@example.", "b@example.")
    assert sample.sampled_count == 2
    assert sample.non_null_count == 2
    assert sample.max_value_length == 10
    assert sample.truncated
    assert "a@example." not in repr(sample)
    assert "values=" not in repr(sample)


def test_csv_file_adapter_returns_empty_sample_for_zero_limit(tmp_path: Path) -> None:
    csv_path = tmp_path / "clientes.csv"
    csv_path.write_text("email\na@example.com\n", encoding="utf-8")
    context = file_context(csv_path, extension=".csv", mime_type="text/csv")
    adapter = build_file_source_adapter(context)
    table = next(iter(adapter.iter_tables()))
    column = next(iter(adapter.iter_columns(table)))

    sample = adapter.get_column_sample(table, column, limit=0)

    assert sample.values == ()
    assert sample.sampled_count == 0
    assert sample.non_null_count == 0
    assert not sample.truncated


def test_csv_file_adapter_rejects_negative_sample_limit(tmp_path: Path) -> None:
    csv_path = tmp_path / "clientes.csv"
    csv_path.write_text("email\na@example.com\n", encoding="utf-8")
    context = file_context(csv_path, extension=".csv", mime_type="text/csv")
    adapter = build_file_source_adapter(context)
    table = next(iter(adapter.iter_tables()))
    column = next(iter(adapter.iter_columns(table)))

    with pytest.raises(ValueError, match="limit must be non-negative"):
        adapter.get_column_sample(table, column, limit=-1)


def test_profile_file_context_profiles_visible_excel_sheets(tmp_path: Path) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise AssertionError("openpyxl must be installed for this test") from exc

    xlsx_path = tmp_path / "reporte.xlsx"
    workbook = Workbook()
    clientes = workbook.active
    clientes.title = "Clientes"
    clientes.append(["email", "rut"])
    clientes.append(["a@example.com", "12.345.678-5"])
    clientes.append([None, None])
    clientes.append(["b@example.com", "11.111.111-1"])

    pagos = workbook.create_sheet("Pagos")
    pagos.append(["card_number", "amount"])
    pagos.append(["4111111111111111", 100])

    oculta = workbook.create_sheet("Oculta")
    oculta.sheet_state = "hidden"
    oculta.append(["secret"])
    oculta.append(["hidden@example.com"])

    workbook.save(xlsx_path)
    workbook.close()

    context = file_context(
        xlsx_path,
        extension=".xlsx",
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    profile = profile_file_context(context)

    assert profile.source_name == "reporte.xlsx"
    assert profile.source_type == "excel"
    assert [table.table_name for table in profile.tables] == ["Clientes", "Pagos"]
    assert [table.row_count for table in profile.tables] == [2, 1]
    assert [column.column_name for column in profile.tables[0].columns] == [
        "email",
        "rut",
    ]
    assert [column.ordinal_position for column in profile.tables[1].columns] == [1, 2]


def test_excel_file_adapter_samples_visible_sheet_column_values(tmp_path: Path) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise AssertionError("openpyxl must be installed for this test") from exc

    xlsx_path = tmp_path / "reporte.xlsx"
    workbook = Workbook()
    clientes = workbook.active
    clientes.title = "Clientes"
    clientes.append(["email", "score"])
    clientes.append(["  a@example.com  ", 100])
    clientes.append([None, None])
    clientes.append(["b@example.com", 200])
    clientes.append(["largo@example.com", 300])

    oculta = workbook.create_sheet("Oculta")
    oculta.sheet_state = "hidden"
    oculta.append(["email"])
    oculta.append(["hidden@example.com"])

    workbook.save(xlsx_path)
    workbook.close()

    context = file_context(
        xlsx_path,
        extension=".xlsx",
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    adapter = build_file_source_adapter(context)
    assert isinstance(adapter, SourceAdapter)

    tables = tuple(adapter.iter_tables())
    assert [candidate.table_name for candidate in tables] == ["Clientes"]
    table = tables[0]
    columns = tuple(adapter.iter_columns(table))

    email_sample = adapter.get_column_sample(
        table,
        columns[0],
        limit=2,
        max_value_length=10,
    )
    score_sample = adapter.get_column_sample(table, columns[1], limit=2)

    assert table.table_name == "Clientes"
    assert email_sample.values == ("a@example.", "b@example.")
    assert email_sample.sampled_count == 2
    assert email_sample.non_null_count == 2
    assert email_sample.truncated
    assert score_sample.values == ("100", "200")
    assert "hidden@example.com" not in repr(email_sample)
    assert "a@example." not in repr(email_sample)


def test_default_file_scan_callback_profiles_and_prints_summary(
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
) -> None:
    csv_path = tmp_path / "clientes.csv"
    csv_path.write_text("email,rut\na@example.com,12.345.678-5\n", encoding="utf-8")
    context = file_context(csv_path, extension=".csv", mime_type="text/csv")

    default_file_scan_callback(context)

    output = capsys.readouterr().out
    assert "profiled_file_scan_context" in output
    assert "file_id=file-001" in output
    assert "source_type=csv" in output
    assert "tables=1" in output
    assert "columns=2" in output


def test_database_scan_request_hides_connection_uri_in_repr() -> None:
    request = DatabaseScanRequest(
        connection_uri="postgresql://user:secret@localhost/db",
        source_name="prod_db",
        dialect="postgresql",
    )

    assert "secret" not in repr(request)
    assert "postgresql://" not in repr(request)
    assert "prod_db" in repr(request)
