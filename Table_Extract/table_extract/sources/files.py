from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable

from table_extract.models import ColumnProfile, ColumnSample, TableProfile
from table_extract.runtime.models import (
    GOOGLE_SPREADSHEET_MIME_TYPE,
    FileScanContext,
    normalize_extension,
)


CSV_SOURCE_TYPE = "csv"
EXCEL_SOURCE_TYPE = "excel"
UNKNOWN_DATA_TYPE = "unknown"


class CsvFileSourceAdapter:
    source_type = CSV_SOURCE_TYPE
    dialect = None

    def __init__(self, context: FileScanContext) -> None:
        self._context = context
        self._path = Path(context.local_path)
        self.source_name = context.file_name
        self.source_uri = context.source_uri
        self._table: TableProfile | None = None
        self._columns: tuple[ColumnProfile, ...] | None = None

    def iter_tables(self) -> Iterable[TableProfile]:
        self._load_profile()
        if self._table is None:
            return iter(())
        return iter((self._table,))

    def iter_columns(self, table: TableProfile) -> Iterable[ColumnProfile]:
        self._load_profile()
        return iter(self._columns or ())

    def get_column_sample(
        self,
        table: TableProfile,
        column: ColumnProfile,
        *,
        limit: int = 1000,
        max_value_length: int = 256,
    ) -> ColumnSample:
        _validate_sample_limit(limit)
        values, truncated = _read_csv_sample(
            self._path,
            _column_index(column),
            limit,
            max_value_length,
        )
        return ColumnSample(
            table_name=table.table_name,
            schema_name=table.schema_name,
            column_name=column.column_name,
            values=values,
            sampled_count=len(values),
            non_null_count=len(values),
            max_value_length=max_value_length,
            truncated=truncated,
        )

    def close(self) -> None:
        return None

    def _load_profile(self) -> None:
        if self._table is not None:
            return

        header: list[str] = []
        row_count = 0
        header, row_count = _read_csv_profile(self._path)

        self._columns = _columns_from_header(header)
        self._table = TableProfile(
            table_name=self._context.file_name,
            table_type=CSV_SOURCE_TYPE,
            columns=(),
            row_count=row_count,
        )


class ExcelFileSourceAdapter:
    source_type = EXCEL_SOURCE_TYPE
    dialect = None

    def __init__(self, context: FileScanContext) -> None:
        self._context = context
        self._path = Path(context.local_path)
        self.source_name = context.file_name
        self.source_uri = context.source_uri
        self._workbook = None
        self._tables: tuple[TableProfile, ...] | None = None
        self._columns_by_table: dict[str, tuple[ColumnProfile, ...]] = {}

    def iter_tables(self) -> Iterable[TableProfile]:
        self._load_profile()
        return iter(self._tables or ())

    def iter_columns(self, table: TableProfile) -> Iterable[ColumnProfile]:
        self._load_profile()
        return iter(self._columns_by_table.get(table.table_name, ()))

    def get_column_sample(
        self,
        table: TableProfile,
        column: ColumnProfile,
        *,
        limit: int = 1000,
        max_value_length: int = 256,
    ) -> ColumnSample:
        _validate_sample_limit(limit)
        self._load_profile()
        worksheet = self._visible_worksheet(table.table_name)
        values, truncated = _read_excel_sample(
            worksheet,
            _column_index(column),
            limit,
            max_value_length,
        )
        return ColumnSample(
            table_name=table.table_name,
            schema_name=table.schema_name,
            column_name=column.column_name,
            values=values,
            sampled_count=len(values),
            non_null_count=len(values),
            max_value_length=max_value_length,
            truncated=truncated,
        )

    def close(self) -> None:
        if self._workbook is not None:
            self._workbook.close()
            self._workbook = None

    def _load_profile(self) -> None:
        if self._tables is not None:
            return

        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "Missing dependency: install openpyxl with "
                "`python -m pip install -r requirements.txt`."
            ) from exc

        self._workbook = load_workbook(
            self._path,
            read_only=True,
            data_only=True,
        )
        tables: list[TableProfile] = []
        columns_by_table: dict[str, tuple[ColumnProfile, ...]] = {}

        for worksheet in self._workbook.worksheets:
            if worksheet.sheet_state != "visible":
                continue

            header = _excel_header(worksheet)
            columns = _columns_from_header(header)
            row_count = _excel_data_row_count(worksheet)
            table = TableProfile(
                table_name=worksheet.title,
                table_type="sheet",
                columns=(),
                row_count=row_count,
            )
            tables.append(table)
            columns_by_table[table.table_name] = columns

        self._tables = tuple(tables)
        self._columns_by_table = columns_by_table

    def _visible_worksheet(self, table_name: str):
        if self._workbook is None:
            raise RuntimeError("Workbook is not loaded.")
        worksheet = self._workbook[table_name]
        if worksheet.sheet_state != "visible":
            raise ValueError(f"Worksheet is not visible: {table_name}")
        return worksheet


def build_file_source_adapter(context: FileScanContext):
    extension = normalize_extension(context.extension or Path(context.local_path).suffix)
    local_extension = normalize_extension(Path(context.local_path).suffix)
    mime_type = (context.mime_type or "").strip().casefold()

    if extension == ".csv" or local_extension == ".csv" or mime_type == "text/csv":
        return CsvFileSourceAdapter(context)
    if (
        extension in {".xlsx", ".xlsm"}
        or local_extension in {".xlsx", ".xlsm"}
        or mime_type == GOOGLE_SPREADSHEET_MIME_TYPE
    ):
        return ExcelFileSourceAdapter(context)

    raise ValueError(
        "Unsupported file type for profiling: "
        f"extension={extension or local_extension} mime_type={context.mime_type}"
    )


def _read_csv_profile(path: Path) -> tuple[list[str], int]:
    last_error: UnicodeDecodeError | None = None
    for encoding in ("utf-8-sig", "latin-1"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                reader = csv.reader(handle)
                try:
                    header = next(reader)
                except StopIteration:
                    return [], 0

                row_count = sum(1 for row in reader if _has_data(row))
                return header, row_count
        except UnicodeDecodeError as exc:
            last_error = exc
            continue
    if last_error:
        raise last_error
    return [], 0


def _read_csv_sample(
    path: Path,
    column_index: int,
    limit: int,
    max_value_length: int,
) -> tuple[tuple[str, ...], bool]:
    last_error: UnicodeDecodeError | None = None
    for encoding in ("utf-8-sig", "latin-1"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                reader = csv.reader(handle)
                next(reader, None)
                return _sample_rows(
                    reader,
                    column_index,
                    limit,
                    max_value_length,
                )
        except UnicodeDecodeError as exc:
            last_error = exc
            continue
    if last_error:
        raise last_error
    return (), False


def _read_excel_sample(
    worksheet,
    column_index: int,
    limit: int,
    max_value_length: int,
) -> tuple[tuple[str, ...], bool]:
    rows = worksheet.iter_rows(min_row=2, values_only=True)
    return _sample_rows(rows, column_index, limit, max_value_length)


def _sample_rows(
    rows,
    column_index: int,
    limit: int,
    max_value_length: int,
) -> tuple[tuple[str, ...], bool]:
    values: list[str] = []
    truncated = False
    if limit == 0:
        return (), False

    for row in rows:
        raw_value = row[column_index] if column_index < len(row) else None
        normalized = _normalize_sample_value(raw_value, max_value_length)
        if normalized is None:
            continue

        value, value_truncated = normalized
        values.append(value)
        truncated = truncated or value_truncated
        if len(values) >= limit:
            break

    return tuple(values), truncated


def _columns_from_header(header: list[object]) -> tuple[ColumnProfile, ...]:
    columns: list[ColumnProfile] = []
    for index, raw_name in enumerate(header, start=1):
        column_name = _normalize_header_value(raw_name, index)
        columns.append(
            ColumnProfile(
                column_name=column_name,
                data_type=UNKNOWN_DATA_TYPE,
                ordinal_position=index,
            )
        )
    return tuple(columns)


def _normalize_header_value(value: object, index: int) -> str:
    if value is None:
        return f"column_{index}"
    text = str(value).strip()
    return text or f"column_{index}"


def _normalize_sample_value(
    value: object,
    max_value_length: int,
) -> tuple[str, bool] | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if len(text) > max_value_length:
        return text[:max_value_length], True
    return text, False


def _validate_sample_limit(limit: int) -> None:
    if limit < 0:
        raise ValueError("limit must be non-negative")


def _column_index(column: ColumnProfile) -> int:
    if column.ordinal_position is None:
        raise ValueError("column.ordinal_position is required for file sampling")
    return column.ordinal_position - 1


def _has_data(row: list[object] | tuple[object, ...]) -> bool:
    return any(value is not None and str(value).strip() for value in row)


def _excel_header(worksheet) -> list[object]:
    for row in worksheet.iter_rows(min_row=1, max_row=1, values_only=True):
        return list(row)
    return []


def _excel_data_row_count(worksheet) -> int:
    row_count = 0
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if _has_data(row):
            row_count += 1
    return row_count
